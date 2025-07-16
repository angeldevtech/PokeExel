import json
from openpyxl import load_workbook

# Load workbook
wb = load_workbook("Pokeexel.xlsx", data_only=True)

# Main data sheet
main_sheet = wb["Lista Pokémon"]

# Region sheet names to check
regions = [
    "Kanto", "Johto", "Hoenn", "Sinnoh", "Teselia",
    "Kalos", "Alola", "Galar", "Paldea"
]

# Preload all region data
region_data = {}
for region in regions:
    sheet = wb[region]
    names = set()
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        cell = row[0]
        if cell and isinstance(cell, str):
            names.add(cell.strip())
    region_data[region.capitalize()] = names  # capitalize to match output

# Start reading from row 4
pokemon_list = []
row = 4
while True:
    row_data = main_sheet[row]

    pokedex_num = row_data[0].value
    if pokedex_num is None:
        break  # Reached end of data

    name = row_data[1].value
    base_exp = row_data[2].value

    ev_yield = {
        "hp": row_data[3].value or 0,
        "attack": row_data[4].value or 0,
        "defense": row_data[5].value or 0,
        "special_attack": row_data[6].value or 0,
        "special_defense": row_data[7].value or 0,
        "speed": row_data[8].value or 0,
    }

    # Check region presence by looking up name in region_data
    found_regions = []
    for region, names in region_data.items():
        if name in names:
            found_regions.append(region)

    pokemon_entry = {
        "id": pokedex_num,
        "name": name,
        "base_exp": base_exp,
        "ev_yield": ev_yield,
        "regions": found_regions
    }

    pokemon_list.append(pokemon_entry)
    row += 1

# Save as JSON
with open("pokemon.json", "w", encoding="utf-8") as f:
    json.dump(pokemon_list, f, ensure_ascii=False, indent=2)
